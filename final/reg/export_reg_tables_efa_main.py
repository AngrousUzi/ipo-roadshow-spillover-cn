"""
Combined EFA output table — selected sheets only.

Sheets written:
  E1_End_EFA    — every-session after-end CAR, stacked top-rivals
  E2_Start_EFA  — every-session after-start CAR, stacked top-rivals
  E3_Before_EFA — every-session before-start CAR (placebo)
  L1_End_a_pros — mod: a_pros_qa × End CAR
  L1_Str_a_pros — mod: a_pros_qa × Start CAR
  L1_Pre_a_pros — mod: a_pros_qa × Before CAR
  M3_FirstDay   — mean firstday regression (Retnfstd, online subs)
  M4_QA         — mean QA regression (structural outcomes)

Output: final/reg/reg_tables_efa_main.xlsx
"""

import os
import pandas as pd
import numpy as np
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

REG_DIR  = Path("final/reg")
OUT_FILE = Path("final/reg/reg_tables_efa_main.xlsx")

MAX_N = 9  # EFA parallel analysis yielded 9 factors; cap display at 9

# ─── File stem fragments ──────────────────────────────────────────────────────
_E = "every_after_end_rc1_rc2_ic_pc_yfe_ife_pltfe_w99_efa"
_S = "every_after_start_rc1_rc2_ic_pc_yfe_ife_pltfe_w99_efa"
_B = "every_before_start_rc1_rc2_ic_pc_yfe_ife_pltfe_w99_efa"

TOP_VARIANTS = [
    ("",      "All Rivals"),
    ("top1",  "Top 1 Rival"),
    ("top3",  "Top 3 Rivals"),
    ("top5",  "Top 5 Rivals"),
    ("top10", "Top 10 Rivals"),
]
TOP_DARK  = ["2E4057", "17A398", "C84B31", "4A4E69", "457B9D"]
TOP_LIGHT = ["D6E4F0", "D5F0EE", "FAEAE5", "DEDCE9", "D6EBF4"]

# ─── Significance helpers ─────────────────────────────────────────────────────
def stars(p):
    if pd.isna(p): return ""
    if p < 0.01:   return "***"
    if p < 0.05:   return "**"
    if p < 0.10:   return "*"
    return ""

def fmt_coef(coef, p):
    if pd.isna(coef): return ""
    return f"{coef:.4f}{stars(p)}"

def fmt_tstat(t):
    if pd.isna(t): return ""
    return f"({t:.3f})"

# ─── Label maps ───────────────────────────────────────────────────────────────
Y_LABEL_MAP = {
    "car_after_end_30min_with925_est1":   "End30m\nw925 e1",
    "car_after_end_30min_with925_est2":   "End30m\nw925 e2",
    "car_after_end_30min_with925_est3":   "End30m\nw925 e3",
    "car_after_end_30min_no925_est1":     "End30m\nno925 e1",
    "car_after_end_30min_no925_est2":     "End30m\nno925 e2",
    "car_after_end_30min_no925_est3":     "End30m\nno925 e3",
    "car_after_end_1hr_with925_est1":     "End1hr\nw925 e1",
    "car_after_end_1hr_with925_est2":     "End1hr\nw925 e2",
    "car_after_end_1hr_with925_est3":     "End1hr\nw925 e3",
    "car_after_end_1hr_no925_est1":       "End1hr\nno925 e1",
    "car_after_end_1hr_no925_est2":       "End1hr\nno925 e2",
    "car_after_end_1hr_no925_est3":       "End1hr\nno925 e3",
    "car_after_start_30min_with925_est1": "Str30m\nw925 e1",
    "car_after_start_30min_with925_est2": "Str30m\nw925 e2",
    "car_after_start_30min_with925_est3": "Str30m\nw925 e3",
    "car_after_start_30min_no925_est1":   "Str30m\nno925 e1",
    "car_after_start_30min_no925_est2":   "Str30m\nno925 e2",
    "car_after_start_30min_no925_est3":   "Str30m\nno925 e3",
    "car_after_start_1hr_with925_est1":   "Str1hr\nw925 e1",
    "car_after_start_1hr_with925_est2":   "Str1hr\nw925 e2",
    "car_after_start_1hr_with925_est3":   "Str1hr\nw925 e3",
    "car_after_start_1hr_no925_est1":     "Str1hr\nno925 e1",
    "car_after_start_1hr_no925_est2":     "Str1hr\nno925 e2",
    "car_after_start_1hr_no925_est3":     "Str1hr\nno925 e3",
    "car_before_start_30min_with925_est1":"Pre30m\nw925 e1",
    "car_before_start_30min_with925_est2":"Pre30m\nw925 e2",
    "car_before_start_30min_with925_est3":"Pre30m\nw925 e3",
    "car_before_start_30min_no925_est1":  "Pre30m\nno925 e1",
    "car_before_start_30min_no925_est2":  "Pre30m\nno925 e2",
    "car_before_start_30min_no925_est3":  "Pre30m\nno925 e3",
    "car_before_start_1hr_with925_est1":  "Pre1hr\nw925 e1",
    "car_before_start_1hr_with925_est2":  "Pre1hr\nw925 e2",
    "car_before_start_1hr_with925_est3":  "Pre1hr\nw925 e3",
    "car_before_start_1hr_no925_est1":    "Pre1hr\nno925 e1",
    "car_before_start_1hr_no925_est2":    "Pre1hr\nno925 e2",
    "car_before_start_1hr_no925_est3":    "Pre1hr\nno925 e3",
    "Retnfstd":              "1st-Day\nReturn",
    "ipo_online_subs_ratio": "Online\nSubs",
    "qa_pairs":              "QA Pairs",
    "speech_count":          "Speech Ct",
    "avg_q_len":             "Avg Q Len",
    "avg_a_len":             "Avg A Len",
    "a_q_len_ratio":         "A/Q Len",
    "num_ratio_in_answer":   "Num Ratio\nAns",
    "n_unique_questioners":  "N Unique\nQ'ers",
}

# EFA uses f1, f2, ... instead of pc1, pc2, ...
X_LABEL_MAP = {
    **{f"f{i}":  f"F{i}"  for i in range(1, 33)},
}

Y_ORDER_END = [
    "car_after_end_30min_with925_est1", "car_after_end_30min_with925_est2", "car_after_end_30min_with925_est3",
    "car_after_end_30min_no925_est1",   "car_after_end_30min_no925_est2",   "car_after_end_30min_no925_est3",
    "car_after_end_1hr_with925_est1",   "car_after_end_1hr_with925_est2",   "car_after_end_1hr_with925_est3",
    "car_after_end_1hr_no925_est1",     "car_after_end_1hr_no925_est2",     "car_after_end_1hr_no925_est3",
]
Y_ORDER_START = [
    "car_after_start_30min_with925_est1","car_after_start_30min_with925_est2","car_after_start_30min_with925_est3",
    "car_after_start_30min_no925_est1",  "car_after_start_30min_no925_est2",  "car_after_start_30min_no925_est3",
    "car_after_start_1hr_with925_est1",  "car_after_start_1hr_with925_est2",  "car_after_start_1hr_with925_est3",
    "car_after_start_1hr_no925_est1",    "car_after_start_1hr_no925_est2",    "car_after_start_1hr_no925_est3",
]
Y_ORDER_BEFORE = [
    "car_before_start_30min_with925_est1","car_before_start_30min_with925_est2","car_before_start_30min_with925_est3",
    "car_before_start_30min_no925_est1",  "car_before_start_30min_no925_est2",  "car_before_start_30min_no925_est3",
    "car_before_start_1hr_with925_est1",  "car_before_start_1hr_with925_est2",  "car_before_start_1hr_with925_est3",
    "car_before_start_1hr_no925_est1",    "car_before_start_1hr_no925_est2",    "car_before_start_1hr_no925_est3",
]
Y_ORDER_FIRSTDAY = ["Retnfstd", "ipo_online_subs_ratio"]
Y_ORDER_QA = [
    "qa_pairs", "speech_count", "avg_q_len", "avg_a_len",
    "a_q_len_ratio", "num_ratio_in_answer", "n_unique_questioners",
]

# ─── Styles ───────────────────────────────────────────────────────────────────
GRP_DARK  = {"am": "1F4E79", "pm": "375623", "all": "833C00"}
GRP_MID   = {"am": "2E75B6", "pm": "548235", "all": "C55A11"}
N_LIGHT   = ["D9E1F2", "E2EFDA", "FCE4D6", "E8D9F0", "DDEBF7",
             "FFF3CD", "E8F5E9", "FCE4EC", "E3F2FD"]  # up to 9

XGRP_FILL = PatternFill("solid", fgColor="F2F2F2")
XGRP_FONT = Font(bold=True, size=9, italic=True)
BOT_FILL  = PatternFill("solid", fgColor="EBF3FB")
BOT_FONT  = Font(bold=True, size=9)
INT_FILL  = PatternFill("solid", fgColor="FFF0F0")
INT_FONT  = Font(size=8, color="C00000")
MOD_FILL  = PatternFill("solid", fgColor="FFF8E7")
MOD_FONT  = Font(size=9, italic=True, bold=True)
THIN = Border(bottom=Side(style="thin"))
MED  = Border(bottom=Side(style="medium"))
C    = Alignment(horizontal="center", vertical="center", wrap_text=True)
L    = Alignment(horizontal="left",   vertical="center", wrap_text=True)

NOTE_EVERY = (
    "Each block groups a different set of top industry rivals.  "
    "Columns: AM/PM | Y outcome | n=1..N factors.  "
    "R²(ctrl): IPO controls + Year FE + Industry FE.  "
    "* p<0.10  ** p<0.05  *** p<0.01   t-stats in parentheses."
)
NOTE_MEAN = (
    "Columns: Level 1 = group (AM/PM), Level 2 = Y outcome, Level 3 = model with n factors.  "
    "Each factor row shows its coefficient from the joint n-factor regression; blank = not in that model.  "
    "R²(ctrl): IPO controls + Year FE + Industry FE.  "
    "* p<0.10  ** p<0.05  *** p<0.01   t-stats in parentheses."
)
NOTE_MOD = (
    "Each block = different top-rival subset.  "
    "Rows: F_i = factor coefficient in moderated+controlled spec; "
    "F_i×MOD = interaction coefficient; MOD = moderator main effect.  "
    "R²(ctrl): IPO controls + Year FE + Industry FE.  "
    "* p<0.10  ** p<0.05  *** p<0.01   t-stats in parentheses."
)

# ─── Data helpers (factor names: f1, f2, ...) ─────────────────────────────────
def melt_efa(df, max_n=MAX_N):
    rows = []
    for n in range(1, max_n + 1):
        sub = df[df["n_pcs"] == n]
        for _, row in sub.iterrows():
            for i in range(1, n + 1):
                rows.append({
                    "group":       row["group"],
                    "n_pcs":       n,
                    "y_col":       row["y_col"],
                    "x_col":       f"f{i}",
                    "coef_ctrl":   row.get(f"coef_f{i}_ctrl",   np.nan),
                    "tstat_ctrl":  row.get(f"tstat_f{i}_ctrl",  np.nan),
                    "pvalue_ctrl": row.get(f"pvalue_f{i}_ctrl", np.nan),
                    "n_obs_ctrl":  row.get("n_obs_ctrl",          np.nan),
                    "r2_ctrl":     row.get("r2_ctrl",             np.nan),
                })
    return pd.DataFrame(rows)


def build_lookup(df_long, n, grp):
    sub = df_long[(df_long["n_pcs"] == n) & (df_long["group"] == grp)]
    return {(r["x_col"], r["y_col"]): r for _, r in sub.iterrows()}


def _build_col_layout(df_long_list, y_cols, groups, max_n):
    all_df = pd.concat(df_long_list, ignore_index=True)
    f_list = [f"f{i}" for i in range(1, max_n + 1)]
    union_lk = {n: {grp: build_lookup(all_df, n, grp) for grp in groups}
                for n in range(1, max_n + 1)}

    def present_ys(grp):
        return [y for y in y_cols
                if any((x, y) in union_lk[n][grp]
                       for n in range(1, max_n + 1) for x in f_list)]

    col = 2
    grp_col_start = {}; ycol_col_start = {}; cell_col = {}; grp_ys = {}
    for grp in groups:
        p_ys = present_ys(grp)
        grp_ys[grp] = p_ys
        if not p_ys:
            continue
        grp_col_start[grp] = col
        for y in p_ys:
            ycol_col_start[(grp, y)] = col
            for n in range(1, max_n + 1):
                cell_col[(grp, y, n)] = col
                col += 1
    return grp_col_start, ycol_col_start, cell_col, grp_ys, col - 2


def _write_triple_header(ws, y_cols, groups, grp_ys, grp_col_start,
                         ycol_col_start, cell_col, max_n):
    corner_fill = PatternFill("solid", fgColor="1F4E79")
    for r_ in range(1, 4):
        ws.cell(r_, 1).fill = corner_fill
    ws.cell(3, 1, "X \\ Y").font = Font(color="FFFFFF", bold=True, size=9)
    ws.cell(3, 1).alignment = C
    ws.merge_cells(start_row=1, start_column=1, end_row=3, end_column=1)

    for grp in groups:
        p_ys = grp_ys[grp]
        if not p_ys:
            continue
        sc = grp_col_start[grp]
        ec = sc + len(p_ys) * max_n - 1
        c = ws.cell(1, sc, grp.upper())
        c.fill = PatternFill("solid", fgColor=GRP_DARK[grp])
        c.font = Font(color="FFFFFF", bold=True, size=11)
        c.alignment = C
        if ec > sc:
            ws.merge_cells(start_row=1, start_column=sc, end_row=1, end_column=ec)

        for y in p_ys:
            sc2 = ycol_col_start[(grp, y)]
            ec2 = sc2 + max_n - 1
            c = ws.cell(2, sc2, Y_LABEL_MAP.get(y, y))
            c.fill = PatternFill("solid", fgColor=GRP_MID[grp])
            c.font = Font(color="FFFFFF", bold=True, size=8)
            c.alignment = C
            if ec2 > sc2:
                ws.merge_cells(start_row=2, start_column=sc2, end_row=2, end_column=ec2)

        for y in p_ys:
            for ni, n in enumerate(range(1, max_n + 1)):
                if (grp, y, n) not in cell_col:
                    continue
                c = ws.cell(3, cell_col[(grp, y, n)], f"n={n}")
                c.fill = PatternFill("solid", fgColor=N_LIGHT[ni % len(N_LIGHT)])
                c.font = Font(bold=True, size=8)
                c.alignment = C

    ws.row_dimensions[1].height = 18
    ws.row_dimensions[2].height = 32
    ws.row_dimensions[3].height = 14


# ─── Every-session sheet writer (stacked top-rival blocks) ───────────────────
def write_sheet_stacked(ws, variant_dfs, y_cols, max_n, sheet_title):
    ws.title = sheet_title[:31]
    groups = ["am", "pm", "all"]
    f_list = [f"f{i}" for i in range(1, max_n + 1)]

    non_empty = [d for d, _, _ in variant_dfs if not d.empty]
    if not non_empty:
        ws.cell(1, 1, "No data (run regressions first)").font = Font(italic=True)
        return

    (grp_col_start, ycol_col_start, cell_col,
     grp_ys, total_data_cols) = _build_col_layout(non_empty, y_cols, groups, max_n)

    _write_triple_header(ws, y_cols, groups, grp_ys, grp_col_start,
                         ycol_col_start, cell_col, max_n)
    row_idx = 4

    for df_long, blk_label, vi in variant_dfs:
        dark_hex  = TOP_DARK[vi % len(TOP_DARK)]
        light_hex = TOP_LIGHT[vi % len(TOP_LIGHT)]
        lookups = ({n: {grp: build_lookup(df_long, n, grp) for grp in groups}
                    for n in range(1, max_n + 1)}
                   if not df_long.empty
                   else {n: {grp: {} for grp in groups} for n in range(1, max_n + 1)})

        sep_fill = PatternFill("solid", fgColor=dark_hex)
        c = ws.cell(row_idx, 1, blk_label)
        c.fill = sep_fill; c.font = Font(color="FFFFFF", bold=True, size=10); c.alignment = L
        for dc in range(total_data_cols):
            ws.cell(row_idx, 2 + dc).fill = sep_fill
        ws.row_dimensions[row_idx].height = 16
        row_idx += 1

        for x in f_list:
            ws.cell(row_idx,     1, X_LABEL_MAP.get(x, x)).alignment = L
            ws.cell(row_idx,     1).font = Font(size=9)
            ws.cell(row_idx + 1, 1, "")
            for grp in groups:
                for y in grp_ys[grp]:
                    for ni, n in enumerate(range(1, max_n + 1)):
                        if (grp, y, n) not in cell_col:
                            continue
                        col_idx = cell_col[(grp, y, n)]
                        r = lookups[n][grp].get((x, y))
                        cell = ws.cell(row_idx, col_idx,
                                       fmt_coef(r.get("coef_ctrl", np.nan),
                                                r.get("pvalue_ctrl", np.nan))
                                       if r is not None else "")
                        cell.alignment = C; cell.font = Font(size=9)
                        tc = ws.cell(row_idx + 1, col_idx,
                                     fmt_tstat(r.get("tstat_ctrl", np.nan))
                                     if r is not None else "")
                        tc.alignment = C
                        tc.font = Font(size=8, color="595959")
                        tc.fill = PatternFill("solid", fgColor=light_hex)
            row_idx += 2

        c = ws.cell(row_idx, 1, "N")
        c.font = BOT_FONT; c.fill = BOT_FILL; c.alignment = L; c.border = THIN
        for grp in groups:
            for y in grp_ys[grp]:
                for n in range(1, max_n + 1):
                    if (grp, y, n) not in cell_col:
                        continue
                    n_val = ""
                    for x in f_list:
                        r = lookups[n][grp].get((x, y))
                        if r is not None:
                            nv = r.get("n_obs_ctrl", np.nan)
                            if not pd.isna(nv):
                                n_val = int(nv); break
                    cell = ws.cell(row_idx, cell_col[(grp, y, n)], n_val)
                    cell.alignment = C; cell.font = BOT_FONT
                    cell.fill = BOT_FILL; cell.border = THIN
        row_idx += 1

        c = ws.cell(row_idx, 1, "R²(ctrl)")
        c.font = BOT_FONT; c.fill = BOT_FILL; c.alignment = L; c.border = MED
        for grp in groups:
            for y in grp_ys[grp]:
                for n in range(1, max_n + 1):
                    if (grp, y, n) not in cell_col:
                        continue
                    rv = ""
                    for x in f_list:
                        r = lookups[n][grp].get((x, y))
                        if r is not None:
                            v = r.get("r2_ctrl", np.nan)
                            if not pd.isna(v):
                                rv = f"{v:.3f}"; break
                    cell = ws.cell(row_idx, cell_col[(grp, y, n)], rv)
                    cell.alignment = C; cell.font = BOT_FONT
                    cell.fill = BOT_FILL; cell.border = MED
        row_idx += 1

    c = ws.cell(row_idx, 1, NOTE_EVERY)
    c.font = Font(size=8, italic=True, color="595959"); c.alignment = L
    if total_data_cols > 1:
        ws.merge_cells(start_row=row_idx, start_column=1,
                       end_row=row_idx, end_column=1 + total_data_cols)
    ws.column_dimensions[get_column_letter(1)].width = 14
    for dc in range(total_data_cols):
        ws.column_dimensions[get_column_letter(2 + dc)].width = 8


# ─── Mean-regression sheet writer ────────────────────────────────────────────
def melt_mean_efa(df, max_n=MAX_N):
    rows = []
    for n in range(1, max_n + 1):
        sub = df[df["n_pcs"] == n]
        for _, row in sub.iterrows():
            for i in range(1, n + 1):
                rows.append({
                    "group":       row["group"],
                    "n_pcs":       n,
                    "y_col":       row["y_col"],
                    "x_col":       f"f{i}",
                    "coef_ctrl":   row.get(f"coef_f{i}_ctrl",   np.nan),
                    "tstat_ctrl":  row.get(f"tstat_f{i}_ctrl",  np.nan),
                    "pvalue_ctrl": row.get(f"pvalue_f{i}_ctrl", np.nan),
                    "n_obs_ctrl":  row.get("n_obs_ctrl",          np.nan),
                    "r2_ctrl":     row.get("r2_ctrl",             np.nan),
                })
    return pd.DataFrame(rows)


def write_sheet_mean(ws, df_long, y_cols, max_n, sheet_title):
    ws.title = sheet_title[:31]
    groups = ["am", "pm", "all"]
    f_list = [f"f{i}" for i in range(1, max_n + 1)]

    lookups = {
        n: {grp: {(r["x_col"], r["y_col"]): r
                  for _, r in df_long[(df_long["n_pcs"] == n) & (df_long["group"] == grp)].iterrows()}
            for grp in groups}
        for n in range(1, max_n + 1)
    }

    def present_ys(grp):
        return [y for y in y_cols
                if any((x, y) in lookups[n][grp] for n in range(1, max_n + 1) for x in f_list)]

    col = 2
    grp_col_start = {}; ycol_col_start = {}; cell_col = {}; grp_ys = {}
    for grp in groups:
        p_ys = present_ys(grp)
        grp_ys[grp] = p_ys
        if not p_ys:
            continue
        grp_col_start[grp] = col
        for y in p_ys:
            ycol_col_start[(grp, y)] = col
            for n in range(1, max_n + 1):
                cell_col[(grp, y, n)] = col
                col += 1

    total_data_cols = col - 2

    corner_fill = PatternFill("solid", fgColor="1F4E79")
    for r_ in range(1, 4):
        ws.cell(r_, 1).fill = corner_fill
    ws.cell(3, 1, "X \\ Y").font = Font(color="FFFFFF", bold=True, size=9)
    ws.cell(3, 1).alignment = C
    ws.merge_cells(start_row=1, start_column=1, end_row=3, end_column=1)

    for grp in groups:
        p_ys = grp_ys[grp]
        if not p_ys:
            continue
        sc = grp_col_start[grp]
        ec = sc + len(p_ys) * max_n - 1
        c = ws.cell(1, sc, grp.upper())
        c.fill = PatternFill("solid", fgColor=GRP_DARK[grp])
        c.font = Font(color="FFFFFF", bold=True, size=11)
        c.alignment = C
        if ec > sc:
            ws.merge_cells(start_row=1, start_column=sc, end_row=1, end_column=ec)

        for y in p_ys:
            sc2 = ycol_col_start[(grp, y)]
            ec2 = sc2 + max_n - 1
            c = ws.cell(2, sc2, Y_LABEL_MAP.get(y, y))
            c.fill = PatternFill("solid", fgColor=GRP_MID[grp])
            c.font = Font(color="FFFFFF", bold=True, size=8)
            c.alignment = C
            if ec2 > sc2:
                ws.merge_cells(start_row=2, start_column=sc2, end_row=2, end_column=ec2)

        for y in p_ys:
            for ni, n in enumerate(range(1, max_n + 1)):
                if (grp, y, n) not in cell_col:
                    continue
                c = ws.cell(3, cell_col[(grp, y, n)], f"n={n}")
                c.fill = PatternFill("solid", fgColor=N_LIGHT[ni % len(N_LIGHT)])
                c.font = Font(bold=True, size=8)
                c.alignment = C

    ws.row_dimensions[1].height = 18
    ws.row_dimensions[2].height = 32
    ws.row_dimensions[3].height = 14

    row_idx = 4
    c = ws.cell(row_idx, 1, "-- EFA --")
    c.font = XGRP_FONT; c.fill = XGRP_FILL; c.alignment = L
    for dc in range(total_data_cols):
        ws.cell(row_idx, 2 + dc).fill = XGRP_FILL
    row_idx += 1

    for x in f_list:
        ws.cell(row_idx,     1, X_LABEL_MAP.get(x, x)).alignment = L
        ws.cell(row_idx,     1).font = Font(size=9)
        ws.cell(row_idx + 1, 1, "")
        for grp in groups:
            for y in grp_ys[grp]:
                for ni, n in enumerate(range(1, max_n + 1)):
                    if (grp, y, n) not in cell_col:
                        continue
                    col_idx = cell_col[(grp, y, n)]
                    r = lookups[n][grp].get((x, y))
                    cell = ws.cell(row_idx, col_idx,
                                   fmt_coef(r.get("coef_ctrl", np.nan),
                                            r.get("pvalue_ctrl", np.nan))
                                   if r is not None else "")
                    cell.alignment = C; cell.font = Font(size=9)
                    tc = ws.cell(row_idx + 1, col_idx,
                                 fmt_tstat(r.get("tstat_ctrl", np.nan))
                                 if r is not None else "")
                    tc.alignment = C
                    tc.font = Font(size=8, color="595959")
                    tc.fill = PatternFill("solid", fgColor=N_LIGHT[ni % len(N_LIGHT)])
        row_idx += 2

    c = ws.cell(row_idx, 1, "N")
    c.font = BOT_FONT; c.fill = BOT_FILL; c.alignment = L; c.border = THIN
    for grp in groups:
        for y in grp_ys[grp]:
            for n in range(1, max_n + 1):
                if (grp, y, n) not in cell_col:
                    continue
                n_val = ""
                for x in f_list:
                    r = lookups[n][grp].get((x, y))
                    if r is not None:
                        nv = r.get("n_obs_ctrl", np.nan)
                        if not pd.isna(nv):
                            n_val = int(nv); break
                cell = ws.cell(row_idx, cell_col[(grp, y, n)], n_val)
                cell.alignment = C; cell.font = BOT_FONT
                cell.fill = BOT_FILL; cell.border = THIN
    row_idx += 1

    c = ws.cell(row_idx, 1, "R²(ctrl)")
    c.font = BOT_FONT; c.fill = BOT_FILL; c.alignment = L; c.border = MED
    for grp in groups:
        for y in grp_ys[grp]:
            for n in range(1, max_n + 1):
                if (grp, y, n) not in cell_col:
                    continue
                rv = ""
                for x in f_list:
                    r = lookups[n][grp].get((x, y))
                    if r is not None:
                        v = r.get("r2_ctrl", np.nan)
                        if not pd.isna(v):
                            rv = f"{v:.3f}"; break
                cell = ws.cell(row_idx, cell_col[(grp, y, n)], rv)
                cell.alignment = C; cell.font = BOT_FONT
                cell.fill = BOT_FILL; cell.border = MED
    row_idx += 1

    c = ws.cell(row_idx, 1, NOTE_MEAN)
    c.font = Font(size=8, italic=True, color="595959"); c.alignment = L
    if total_data_cols > 1:
        ws.merge_cells(start_row=row_idx, start_column=1,
                       end_row=row_idx, end_column=1 + total_data_cols)
    ws.column_dimensions[get_column_letter(1)].width = 12
    for dc in range(total_data_cols):
        ws.column_dimensions[get_column_letter(2 + dc)].width = 8


# ─── Mod data helpers ─────────────────────────────────────────────────────────
def melt_efa_mod(df, mod_name, max_n=MAX_N):
    rows = []
    for n in range(1, max_n + 1):
        sub = df[df["n_pcs"] == n]
        for _, row in sub.iterrows():
            for i in range(1, n + 1):
                fi = f"f{i}"
                rows.append({
                    "group":           row["group"],
                    "n_pcs":           n,
                    "y_col":           row["y_col"],
                    "x_col":           fi,
                    "coef_f_mod":      row.get(f"coef_{fi}_mod_{mod_name}_ctrl",   np.nan),
                    "tstat_f_mod":     row.get(f"tstat_{fi}_mod_{mod_name}_ctrl",  np.nan),
                    "pvalue_f_mod":    row.get(f"pvalue_{fi}_mod_{mod_name}_ctrl", np.nan),
                    "coef_interact":   row.get(f"coef_interact_{mod_name}_{fi}_ctrl",   np.nan),
                    "tstat_interact":  row.get(f"tstat_interact_{mod_name}_{fi}_ctrl",  np.nan),
                    "pvalue_interact": row.get(f"pvalue_interact_{mod_name}_{fi}_ctrl", np.nan),
                    "coef_mod":        row.get(f"coef_mod_{mod_name}_ctrl",  np.nan),
                    "pval_mod":        row.get(f"pval_mod_{mod_name}_ctrl",  np.nan),
                    "n_obs_mod":       row.get(f"n_obs_mod_{mod_name}_ctrl", np.nan),
                    "r2_mod":          row.get(f"r2_mod_{mod_name}_ctrl",    np.nan),
                })
    return pd.DataFrame(rows) if rows else pd.DataFrame()


def write_sheet_stacked_mod(ws, variant_dfs_mod, y_cols, max_n,
                             mod_name, mod_display, sheet_title):
    ws.title = sheet_title[:31]
    groups = ["am", "pm", "all"]
    f_list = [f"f{i}" for i in range(1, max_n + 1)]

    non_empty = [d for d, _, _ in variant_dfs_mod if not d.empty]
    if not non_empty:
        ws.cell(1, 1, "No data (run regressions first)").font = Font(italic=True)
        return

    (grp_col_start, ycol_col_start, cell_col,
     grp_ys, total_data_cols) = _build_col_layout(non_empty, y_cols, groups, max_n)

    _write_triple_header(ws, y_cols, groups, grp_ys, grp_col_start,
                         ycol_col_start, cell_col, max_n)
    row_idx = 4

    for df_long_mod, blk_label, vi in variant_dfs_mod:
        dark_hex  = TOP_DARK[vi % len(TOP_DARK)]
        light_hex = TOP_LIGHT[vi % len(TOP_LIGHT)]
        lookups = ({n: {grp: build_lookup(df_long_mod, n, grp) for grp in groups}
                    for n in range(1, max_n + 1)}
                   if not df_long_mod.empty
                   else {n: {grp: {} for grp in groups} for n in range(1, max_n + 1)})

        sep_fill = PatternFill("solid", fgColor=dark_hex)
        c = ws.cell(row_idx, 1, blk_label)
        c.fill = sep_fill; c.font = Font(color="FFFFFF", bold=True, size=10); c.alignment = L
        for dc in range(total_data_cols):
            ws.cell(row_idx, 2 + dc).fill = sep_fill
        ws.row_dimensions[row_idx].height = 16
        row_idx += 1

        for x in f_list:
            x_lbl = X_LABEL_MAP.get(x, x)
            ws.cell(row_idx,     1, x_lbl).alignment = L
            ws.cell(row_idx,     1).font = Font(size=9)
            ws.cell(row_idx + 1, 1, "")
            ws.cell(row_idx + 2, 1, f"{x_lbl} × {mod_display}").alignment = L
            ws.cell(row_idx + 2, 1).font = Font(size=9, color="C00000")
            ws.cell(row_idx + 3, 1, "")

            for grp in groups:
                for y in grp_ys[grp]:
                    for ni, n in enumerate(range(1, max_n + 1)):
                        if (grp, y, n) not in cell_col:
                            continue
                        col_idx = cell_col[(grp, y, n)]
                        r = lookups[n][grp].get((x, y))

                        cell = ws.cell(row_idx, col_idx,
                                       fmt_coef(r.get("coef_f_mod", np.nan),
                                                r.get("pvalue_f_mod", np.nan))
                                       if r is not None else "")
                        cell.alignment = C; cell.font = Font(size=9)

                        tc = ws.cell(row_idx + 1, col_idx,
                                     fmt_tstat(r.get("tstat_f_mod", np.nan))
                                     if r is not None else "")
                        tc.alignment = C
                        tc.font = Font(size=8, color="595959")
                        tc.fill = PatternFill("solid", fgColor=light_hex)

                        ic = ws.cell(row_idx + 2, col_idx,
                                     fmt_coef(r.get("coef_interact", np.nan),
                                              r.get("pvalue_interact", np.nan))
                                     if r is not None else "")
                        ic.alignment = C; ic.font = Font(size=9, color="C00000")

                        itc = ws.cell(row_idx + 3, col_idx,
                                      fmt_tstat(r.get("tstat_interact", np.nan))
                                      if r is not None else "")
                        itc.alignment = C; itc.font = INT_FONT; itc.fill = INT_FILL
            row_idx += 4

        c = ws.cell(row_idx, 1, f"MOD ({mod_display})")
        c.font = MOD_FONT; c.fill = MOD_FILL; c.alignment = L
        for grp in groups:
            for y in grp_ys[grp]:
                for n in range(1, max_n + 1):
                    if (grp, y, n) not in cell_col:
                        continue
                    col_idx = cell_col[(grp, y, n)]
                    r = lookups[n][grp].get(("f1", y))
                    if r is None:
                        for x in f_list:
                            r = lookups[n][grp].get((x, y))
                            if r is not None: break
                    cell = ws.cell(row_idx, col_idx,
                                   fmt_coef(r.get("coef_mod", np.nan),
                                            r.get("pval_mod", np.nan))
                                   if r is not None else "")
                    cell.alignment = C; cell.font = MOD_FONT; cell.fill = MOD_FILL
        row_idx += 1

        c = ws.cell(row_idx, 1, "N")
        c.font = BOT_FONT; c.fill = BOT_FILL; c.alignment = L; c.border = THIN
        for grp in groups:
            for y in grp_ys[grp]:
                for n in range(1, max_n + 1):
                    if (grp, y, n) not in cell_col:
                        continue
                    n_val = ""
                    for x in f_list:
                        r = lookups[n][grp].get((x, y))
                        if r is not None:
                            nv = r.get("n_obs_mod", np.nan)
                            if not pd.isna(nv):
                                n_val = int(nv); break
                    cell = ws.cell(row_idx, cell_col[(grp, y, n)], n_val)
                    cell.alignment = C; cell.font = BOT_FONT
                    cell.fill = BOT_FILL; cell.border = THIN
        row_idx += 1

        c = ws.cell(row_idx, 1, "R²(mod+ctrl)")
        c.font = BOT_FONT; c.fill = BOT_FILL; c.alignment = L; c.border = MED
        for grp in groups:
            for y in grp_ys[grp]:
                for n in range(1, max_n + 1):
                    if (grp, y, n) not in cell_col:
                        continue
                    rv = ""
                    for x in f_list:
                        r = lookups[n][grp].get((x, y))
                        if r is not None:
                            v = r.get("r2_mod", np.nan)
                            if not pd.isna(v):
                                rv = f"{v:.3f}"; break
                    cell = ws.cell(row_idx, cell_col[(grp, y, n)], rv)
                    cell.alignment = C; cell.font = BOT_FONT
                    cell.fill = BOT_FILL; cell.border = MED
        row_idx += 1

    c = ws.cell(row_idx, 1, NOTE_MOD)
    c.font = Font(size=8, italic=True, color="595959"); c.alignment = L
    if total_data_cols > 1:
        ws.merge_cells(start_row=row_idx, start_column=1,
                       end_row=row_idx, end_column=1 + total_data_cols)
    ws.column_dimensions[get_column_letter(1)].width = 18
    for dc in range(total_data_cols):
        ws.column_dimensions[get_column_letter(2 + dc)].width = 8


# ─── File loading helpers ─────────────────────────────────────────────────────
def melt_efa_from_mod(df, mod_name, max_n=MAX_N):
    rows = []
    for n in range(1, max_n + 1):
        sub = df[df["n_pcs"] == n]
        for _, row in sub.iterrows():
            for i in range(1, n + 1):
                fi = f"f{i}"
                rows.append({
                    "group":       row["group"],
                    "n_pcs":       n,
                    "y_col":       row["y_col"],
                    "x_col":       fi,
                    "coef_ctrl":   row.get(f"coef_{fi}_mod_{mod_name}_ctrl",   np.nan),
                    "tstat_ctrl":  row.get(f"tstat_{fi}_mod_{mod_name}_ctrl",  np.nan),
                    "pvalue_ctrl": row.get(f"pvalue_{fi}_mod_{mod_name}_ctrl", np.nan),
                    "n_obs_ctrl":  row.get(f"n_obs_mod_{mod_name}_ctrl",       np.nan),
                    "r2_ctrl":     row.get(f"r2_mod_{mod_name}_ctrl",          np.nan),
                })
    return pd.DataFrame(rows)


def _load_efa(base_frag, top_suffix="", mod_key="a_pros_qa"):
    part = f"_{top_suffix}" if top_suffix else ""
    fname = REG_DIR / f"reg_bivariate_grouped_{base_frag}{part}_{mod_key}.csv"
    try:
        return melt_efa_from_mod(pd.read_csv(fname), mod_key)
    except FileNotFoundError:
        print(f"  WARNING: {fname.name} not found — empty block")
        return pd.DataFrame()


def _load_efa_mod(base_frag, mod_key, top_suffix=""):
    part = f"_{top_suffix}" if top_suffix else ""
    fname = REG_DIR / f"reg_bivariate_grouped_{base_frag}{part}_{mod_key}.csv"
    try:
        return melt_efa_mod(pd.read_csv(fname), mod_key)
    except FileNotFoundError:
        print(f"  WARNING: {fname.name} not found — empty block")
        return pd.DataFrame()


def _load_mean(fname):
    try:
        return melt_mean_efa(pd.read_csv(REG_DIR / fname))
    except FileNotFoundError:
        print(f"  WARNING: {fname} not found — empty")
        return pd.DataFrame()


# ─── Load data ────────────────────────────────────────────────────────────────
print("Loading every-session EFA data...")
dfs_end    = [(_load_efa(_E, sfx), lbl, vi) for vi, (sfx, lbl) in enumerate(TOP_VARIANTS)]
dfs_start  = [(_load_efa(_S, sfx), lbl, vi) for vi, (sfx, lbl) in enumerate(TOP_VARIANTS)]
dfs_before = [(_load_efa(_B, sfx), lbl, vi) for vi, (sfx, lbl) in enumerate(TOP_VARIANTS)]

print("Loading a_pros_qa mod data...")
dfs_mod_end    = [(_load_efa_mod(_E, "a_pros_qa", sfx), lbl, vi) for vi, (sfx, lbl) in enumerate(TOP_VARIANTS)]
dfs_mod_start  = [(_load_efa_mod(_S, "a_pros_qa", sfx), lbl, vi) for vi, (sfx, lbl) in enumerate(TOP_VARIANTS)]
dfs_mod_before = [(_load_efa_mod(_B, "a_pros_qa", sfx), lbl, vi) for vi, (sfx, lbl) in enumerate(TOP_VARIANTS)]

print("Loading mean-regression EFA data...")
df_first = _load_mean("reg_bivariate_grouped_mean_firstday_ic_yfe_ife_efa_pltfe.csv")
df_qa    = _load_mean("reg_bivariate_grouped_mean_qa_ctrl_fe_ife_efa_pltfe.csv")

# ─── Build workbook ───────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
wb.remove(wb.active)

for sheet_name, variant_list, y_order in [
    ("E1_End_EFA",    dfs_end,    Y_ORDER_END),
    ("E2_Start_EFA",  dfs_start,  Y_ORDER_START),
    ("E3_Before_EFA", dfs_before, Y_ORDER_BEFORE),
]:
    ws = wb.create_sheet(sheet_name)
    write_sheet_stacked(ws, variant_list, y_order, MAX_N, sheet_name)
    print(f"  Written: {sheet_name}")

for sheet_name, variant_list, y_order in [
    ("L1_End_a_pros",  dfs_mod_end,    Y_ORDER_END),
    ("L1_Str_a_pros",  dfs_mod_start,  Y_ORDER_START),
    ("L1_Pre_a_pros",  dfs_mod_before, Y_ORDER_BEFORE),
]:
    ws = wb.create_sheet(sheet_name)
    write_sheet_stacked_mod(ws, variant_list, y_order, MAX_N,
                            "a_pros_qa", "A-ProsQA", sheet_name)
    print(f"  Written: {sheet_name}")

for sheet_name, df_, y_order in [
    ("M3_FirstDay", df_first, Y_ORDER_FIRSTDAY),
    ("M4_QA",       df_qa,    Y_ORDER_QA),
]:
    ws = wb.create_sheet(sheet_name)
    write_sheet_mean(ws, df_, y_order, MAX_N, sheet_name)
    print(f"  Written: {sheet_name}")

tmp = OUT_FILE.with_suffix(".tmp.xlsx")
wb.save(tmp)
try:
    tmp.replace(OUT_FILE)
except PermissionError:
    print(f"  WARNING: {OUT_FILE.name} is open in another program — saved as {tmp.name} instead")
    OUT_FILE = tmp
print(f"\nSaved: {OUT_FILE}  ({os.path.getsize(OUT_FILE) // 1024} KB)")
print(f"Total sheets: {len(wb.worksheets)}")
