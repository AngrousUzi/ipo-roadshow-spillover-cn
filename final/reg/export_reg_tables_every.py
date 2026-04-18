"""
Export regression results (every-observation panel) to financial-format Excel tables.
Mirrors export_reg_tables.py but reads the *_every_*_rc1_rc2_ic_pc_yfe_ife_w99* CSVs.
- One sheet per panel; all 4 (session × group) combos shown side-by-side.
- Double-row column header:
    Row 1: session×group block label  (e.g. "推介 AM", merged across Y columns)
    Row 2: individual Y window labels (e.g. "End 30m w925 e1")
- X in first column, rows = coef + t-stat pairs.
- Controlled + FE results only.
- R²(ctrl) = R² of model with controls + Year FE + Industry FE.
"""

import pandas as pd
import numpy as np
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

REG_DIR  = Path("final/reg")
OUT_FILE = Path("final/reg/reg_tables_every.xlsx")

# ─── Session / group ordering ─────────────────────────────────────────────────
SESS_INTRO  = "\u63a8\u4ecb"   # 推介
SESS_THANKS = "\u7b54\u8c22"   # 答谢
ALL_COMBOS  = [
    (SESS_INTRO,  "am", "推介 AM"),
    (SESS_INTRO,  "pm", "推介 PM"),
    (SESS_THANKS, "am", "答谢 AM"),
    (SESS_THANKS, "pm", "答谢 PM"),
]

# ─── Significance stars ───────────────────────────────────────────────────────
def stars(p):
    if pd.isna(p): return ""
    if p < 0.01:   return "***"
    if p < 0.05:   return "**"
    if p < 0.10:   return "*"
    return ""

def fmt_coef(coef, p):
    if pd.isna(coef): return ""
    return f"{coef:.4f}{stars(p)}"

def fmt_tstat(t):
    if pd.isna(t): return ""
    return f"({t:.3f})"

# ─── Label maps ───────────────────────────────────────────────────────────────
Y_LABEL_MAP = {
    "car_after_end_30min_with925_est1":   "End30m\nw925 e1",
    "car_after_end_30min_with925_est2":   "End30m\nw925 e2",
    "car_after_end_30min_with925_est3":   "End30m\nw925 e3",
    "car_after_end_30min_no925_est1":     "End30m\nno925 e1",
    "car_after_end_30min_no925_est2":     "End30m\nno925 e2",
    "car_after_end_30min_no925_est3":     "End30m\nno925 e3",
    "car_after_end_1hr_with925_est1":     "End1hr\nw925 e1",
    "car_after_end_1hr_with925_est2":     "End1hr\nw925 e2",
    "car_after_end_1hr_with925_est3":     "End1hr\nw925 e3",
    "car_after_end_1hr_no925_est1":       "End1hr\nno925 e1",
    "car_after_end_1hr_no925_est2":       "End1hr\nno925 e2",
    "car_after_end_1hr_no925_est3":       "End1hr\nno925 e3",
    "car_after_start_30min_with925_est1": "Str30m\nw925 e1",
    "car_after_start_30min_with925_est2": "Str30m\nw925 e2",
    "car_after_start_30min_with925_est3": "Str30m\nw925 e3",
    "car_after_start_30min_no925_est1":   "Str30m\nno925 e1",
    "car_after_start_30min_no925_est2":   "Str30m\nno925 e2",
    "car_after_start_30min_no925_est3":   "Str30m\nno925 e3",
    "car_after_start_1hr_with925_est1":   "Str1hr\nw925 e1",
    "car_after_start_1hr_with925_est2":   "Str1hr\nw925 e2",
    "car_after_start_1hr_with925_est3":   "Str1hr\nw925 e3",
    "car_after_start_1hr_no925_est1":     "Str1hr\nno925 e1",
    "car_after_start_1hr_no925_est2":     "Str1hr\nno925 e2",
    "car_after_start_1hr_no925_est3":     "Str1hr\nno925 e3",
    "Retnfstd":                           "1st-Day\nReturn",
    "ipo_online_subs_ratio":              "Online\nSubs",
    "qa_pairs":              "QA Pairs",
    "speech_count":          "Speech Ct",
    "avg_q_len":             "Avg Q Len",
    "avg_a_len":             "Avg A Len",
    "a_q_len_ratio":         "A/Q Len",
    "num_ratio_in_answer":   "Num Ratio\nAns",
    "n_unique_questioners":  "N Unique\nQ'ers",
    "q_ann_positive_ratio":  "q_ann_pos",
    "q_ann_negative_ratio":  "q_ann_neg",
    "q_ann_tone_score":      "q_ann_tone",
    "q_social_positive_ratio": "q_soc_pos",
    "q_social_negative_ratio": "q_soc_neg",
    "q_social_tone_score":   "q_soc_tone",
    "q_competition_ratio":   "q_comp",
    "q_prospect_ratio":      "q_prospect",
    "q_policy_pos_ratio":    "q_pol_pos",
    "q_policy_neg_ratio":    "q_pol_neg",
    "q_total_words":         "q_tot_wds",
    "a_ann_positive_ratio":  "a_ann_pos",
    "a_ann_negative_ratio":  "a_ann_neg",
    "a_ann_tone_score":      "a_ann_tone",
    "a_social_positive_ratio": "a_soc_pos",
    "a_social_negative_ratio": "a_soc_neg",
    "a_social_tone_score":   "a_soc_tone",
    "a_competition_ratio":   "a_comp",
    "a_prospect_ratio":      "a_prospect",
    "a_policy_pos_ratio":    "a_pol_pos",
    "a_policy_neg_ratio":    "a_pol_neg",
    "a_total_words":         "a_tot_wds",
}

X_LABEL_MAP = {
    "ann_positive_ratio":     "Ann. Positive",
    "ann_negative_ratio":     "Ann. Negative",
    "ann_tone_score":         "Ann. Tone",
    "social_positive_ratio":  "Social Positive",
    "social_negative_ratio":  "Social Negative",
    "social_tone_score":      "Social Tone",
    "competition_ratio":      "Competition",
    "prospect_ratio":         "Prospect",
    "policy_pos_ratio":       "Policy Pos.",
    "policy_neg_ratio":       "Policy Neg.",
    "total_words":            "Total Words",
    "total_chars":            "Total Chars",
    "f0_mean":                "F0 Mean",
    "f0_std":                 "F0 Std",
    "f0_range":               "F0 Range",
    "f0_slope":               "F0 Slope",
    "rms_mean":               "RMS Mean",
    "rms_std":                "RMS Std",
    "rms_cv":                 "RMS CV",
    "rms_dynamic_range":      "RMS Dyn.Range",
    "rms_snr_proxy":          "RMS SNR Proxy",
    "articulation_rate":      "Articulation Rate",
    "speech_rate":            "Speech Rate",
    "voiced_fraction":        "Voiced Fraction",
    "pause_rate":             "Pause Rate",
    "n_pauses_per_min":       "Pauses/Min",
    "mean_pause_duration":    "Mean Pause Dur",
    "duration_s":             "Duration (s)",
    "asr_logprob_mean":       "ASR Logprob",
    "frames_analyzed":        "Frames Analyzed",
    "frames_with_face":       "Frames w/Face",
    "gaze_at_camera_ratio_5":  "Gaze Cam. 5\u00b0",
    "gaze_at_camera_ratio_10": "Gaze Cam. 10\u00b0",
    "gaze_at_camera_ratio_15": "Gaze Cam. 15\u00b0",
    "gaze_at_camera_ratio_20": "Gaze Cam. 20\u00b0",
    "gaze_x_mean":            "Gaze X Mean",
    "gaze_x_std":             "Gaze X Std",
    "gaze_y_mean":            "Gaze Y Mean",
    "gaze_y_std":             "Gaze Y Std",
    "head_frontal_ratio_5":   "Head Frontal 5\u00b0",
    "head_frontal_ratio_10":  "Head Frontal 10\u00b0",
    "head_frontal_ratio_15":  "Head Frontal 15\u00b0",
    "head_frontal_ratio_20":  "Head Frontal 20\u00b0",
    "head_pitch_mean":        "Head Pitch Mean",
    "head_pitch_std":         "Head Pitch Std",
    "head_yaw_mean":          "Head Yaw Mean",
    "head_yaw_std":           "Head Yaw Std",
    # visual_fer
    "positive_ratio":         "FER Positive",
    "negative_ratio":         "FER Negative",
    "neutral_ratio":          "FER Neutral",
    "net_positive":           "FER Net Positive",
    "emo_angry":              "FER Angry",
    "emo_contempt":           "FER Contempt",
    "emo_disgust":            "FER Disgust",
    "emo_fear":               "FER Fear",
    "emo_happy":              "FER Happy",
    "emo_neutral":            "FER Neutral (emo)",
    "emo_sad":                "FER Sad",
    "emo_surprise":           "FER Surprise",
    "face_detect_rate":       "Face Detect Rate",
}

X_GROUPS = {
    "verbal": [
        "ann_positive_ratio", "ann_negative_ratio", "ann_tone_score",
        "social_positive_ratio", "social_negative_ratio", "social_tone_score",
        "competition_ratio", "prospect_ratio",
        "policy_pos_ratio", "policy_neg_ratio",
        "total_words", "total_chars",
    ],
    "vocal": [
        "f0_mean", "f0_std", "f0_range", "f0_slope",
        "rms_mean", "rms_std", "rms_cv", "rms_dynamic_range", "rms_snr_proxy",
        "articulation_rate", "speech_rate", "voiced_fraction",
        "pause_rate", "n_pauses_per_min", "mean_pause_duration",
        "duration_s", "asr_logprob_mean",
    ],
    "visual": [
        "frames_analyzed", "frames_with_face",
        "gaze_at_camera_ratio_5", "gaze_at_camera_ratio_10",
        "gaze_at_camera_ratio_15", "gaze_at_camera_ratio_20",
        "gaze_x_mean", "gaze_x_std", "gaze_y_mean", "gaze_y_std",
        "head_frontal_ratio_5", "head_frontal_ratio_10",
        "head_frontal_ratio_15", "head_frontal_ratio_20",
        "head_pitch_mean", "head_pitch_std",
        "head_yaw_mean", "head_yaw_std",
    ],
    "visual_fer": [
        "positive_ratio", "negative_ratio", "neutral_ratio", "net_positive",
        "emo_angry", "emo_contempt", "emo_disgust", "emo_fear",
        "emo_happy", "emo_neutral", "emo_sad", "emo_surprise",
        "face_detect_rate",
    ],
}

Y_ORDER_END = [
    "car_after_end_30min_with925_est1", "car_after_end_30min_with925_est2", "car_after_end_30min_with925_est3",
    "car_after_end_30min_no925_est1",   "car_after_end_30min_no925_est2",   "car_after_end_30min_no925_est3",
    "car_after_end_1hr_with925_est1",   "car_after_end_1hr_with925_est2",   "car_after_end_1hr_with925_est3",
    "car_after_end_1hr_no925_est1",     "car_after_end_1hr_no925_est2",     "car_after_end_1hr_no925_est3",
]
Y_ORDER_START = [
    "car_after_start_30min_with925_est1", "car_after_start_30min_with925_est2", "car_after_start_30min_with925_est3",
    "car_after_start_30min_no925_est1",   "car_after_start_30min_no925_est2",   "car_after_start_30min_no925_est3",
    "car_after_start_1hr_with925_est1",   "car_after_start_1hr_with925_est2",   "car_after_start_1hr_with925_est3",
    "car_after_start_1hr_no925_est1",     "car_after_start_1hr_no925_est2",     "car_after_start_1hr_no925_est3",
]

# ─── Styles ───────────────────────────────────────────────────────────────────
COMBO_FILLS = [
    PatternFill("solid", fgColor="1F4E79"),
    PatternFill("solid", fgColor="2E75B6"),
    PatternFill("solid", fgColor="70AD47"),
    PatternFill("solid", fgColor="548235"),
]
COMBO_FONTS = [Font(color="FFFFFF", bold=True, size=9)] * 4
Y_HDR_FILLS = [
    PatternFill("solid", fgColor="BDD7EE"),
    PatternFill("solid", fgColor="DAEEF3"),
    PatternFill("solid", fgColor="E2EFDA"),
    PatternFill("solid", fgColor="C6E0B4"),
]
Y_HDR_FONT = Font(bold=True, size=8)

XGRP_FILL  = PatternFill("solid", fgColor="F2F2F2")
XGRP_FONT  = Font(bold=True, size=9, italic=True)
BOT_FILL   = PatternFill("solid", fgColor="EBF3FB")
BOT_FONT   = Font(bold=True, size=9)
TSTAT_FILLS= [
    PatternFill("solid", fgColor="EAF3FB"),
    PatternFill("solid", fgColor="F0F8FF"),
    PatternFill("solid", fgColor="F0FBF0"),
    PatternFill("solid", fgColor="EAF5E8"),
]
INT_FILL   = PatternFill("solid", fgColor="FFF0F0")
THIN       = Border(bottom=Side(style="thin"))
MED        = Border(bottom=Side(style="medium"))
C          = Alignment(horizontal="center", vertical="center", wrap_text=True)
L          = Alignment(horizontal="left",   vertical="center", wrap_text=True)

NOTE_TEXT = (
    "R\u00b2(ctrl) = R\u00b2 of controlled model: IPO controls "
    "+ Year FE + Industry FE (panel: every session observation, winsorized at 99%).  "
    "* p<0.10  ** p<0.05  *** p<0.01   t-statistics in parentheses."
)


def _build_lookup(df, sess, grp):
    sub = df[(df["session"] == sess) & (df["group"] == grp)]
    return {(r["x_col"], r["y_col"]): r for _, r in sub.iterrows()}

def _build_lookup_suffixed(df, sess, grp, suffix):
    sub = df[(df["session"] == sess) & (df["group"] == grp)]
    return {(r["x_col"], r["y_col"] + suffix): r for _, r in sub.iterrows()}


def _write_double_header(ws, y_cols, all_lookups):
    ws.cell(1, 1, "").fill = PatternFill("solid", fgColor="1F4E79")
    ws.cell(2, 1, "X \\ Y").fill = PatternFill("solid", fgColor="1F4E79")
    ws.cell(2, 1).font = Font(color="FFFFFF", bold=True, size=9)
    ws.cell(2, 1).alignment = C
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)

    col = 2
    start_col_per_combo = {}
    for ci, (lookup, combo_idx, combo_label) in enumerate(all_lookups):
        present = [y for y in y_cols if any((x, y) in lookup for gx in X_GROUPS.values() for x in gx)]
        if not present:
            continue
        start_col_per_combo[ci] = col
        end_col = col + len(present) - 1
        c = ws.cell(1, col, combo_label)
        c.fill = COMBO_FILLS[combo_idx % 4]
        c.font = COMBO_FONTS[combo_idx % 4]
        c.alignment = C
        if end_col > col:
            ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=end_col)
        for y in present:
            c = ws.cell(2, col, Y_LABEL_MAP.get(y, y))
            c.fill = Y_HDR_FILLS[combo_idx % 4]
            c.font = Y_HDR_FONT
            c.alignment = C
            col += 1

    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 36
    return start_col_per_combo, 3


def write_sheet(ws, all_lookups, y_cols, x_groups,
                coef_col, tstat_col, pval_col,
                n_ctrl_col, r2_col, sheet_title):
    ws.title = sheet_title[:31]

    start_col_per_combo, _ = _write_double_header(ws, y_cols, all_lookups)

    row_idx = 3
    for grp_lbl, x_list in x_groups.items():
        active = []
        for x in x_list:
            for lookup, ci, cl in all_lookups:
                if any((x, y) in lookup for y in y_cols):
                    active.append(x)
                    break
        if not active:
            continue

        c = ws.cell(row_idx, 1, f"-- {grp_lbl.upper()} --")
        c.font = XGRP_FONT; c.fill = XGRP_FILL; c.alignment = L
        for ci, (lookup, combo_idx, _) in enumerate(all_lookups):
            if ci in start_col_per_combo:
                sc = start_col_per_combo[ci]
                ny = sum(1 for y in y_cols if any((x, y) in lookup for gx in x_groups.values() for x in gx))
                for dc in range(ny):
                    ws.cell(row_idx, sc + dc).fill = XGRP_FILL
        row_idx += 1

        for x in active:
            x_lbl = X_LABEL_MAP.get(x, x)
            ws.cell(row_idx, 1, x_lbl).alignment = L
            ws.cell(row_idx, 1).font = Font(size=9)
            ws.cell(row_idx + 1, 1, "")

            for ci, (lookup, combo_idx, _) in enumerate(all_lookups):
                if ci not in start_col_per_combo:
                    continue
                present_y = [y for y in y_cols if any((x2, y) in lookup for gx in x_groups.values() for x2 in gx)]
                sc = start_col_per_combo[ci]
                tstat_fill = TSTAT_FILLS[combo_idx % 4]
                for jy, y in enumerate(present_y):
                    r = lookup.get((x, y))
                    col_idx = sc + jy
                    if r is not None:
                        cell = ws.cell(row_idx, col_idx,
                                       fmt_coef(r.get(coef_col, np.nan), r.get(pval_col, np.nan)))
                    else:
                        cell = ws.cell(row_idx, col_idx, "")
                    cell.alignment = C; cell.font = Font(size=9)
                    if r is not None:
                        tc = ws.cell(row_idx + 1, col_idx,
                                     fmt_tstat(r.get(tstat_col, np.nan)))
                    else:
                        tc = ws.cell(row_idx + 1, col_idx, "")
                    tc.alignment = C
                    tc.font = Font(size=8, color="595959")
                    tc.fill = tstat_fill
            row_idx += 2

    # N row
    c = ws.cell(row_idx, 1, "N")
    c.font = BOT_FONT; c.fill = BOT_FILL; c.alignment = L; c.border = THIN
    for ci, (lookup, combo_idx, _) in enumerate(all_lookups):
        if ci not in start_col_per_combo:
            continue
        present_y = [y for y in y_cols if any((x, y) in lookup for gx in x_groups.values() for x in gx)]
        sc = start_col_per_combo[ci]
        for jy, y in enumerate(present_y):
            n_val = ""
            for gx in x_groups.values():
                for x in gx:
                    r = lookup.get((x, y))
                    if r is not None:
                        nv = r.get(n_ctrl_col, np.nan)
                        if not pd.isna(nv):
                            n_val = int(nv); break
                if n_val != "": break
            cell = ws.cell(row_idx, sc + jy, n_val)
            cell.alignment = C; cell.font = BOT_FONT
            cell.fill = BOT_FILL; cell.border = THIN
    row_idx += 1

    # R² row
    c = ws.cell(row_idx, 1, "R\u00b2(ctrl)")
    c.font = BOT_FONT; c.fill = BOT_FILL; c.alignment = L; c.border = MED
    for ci, (lookup, combo_idx, _) in enumerate(all_lookups):
        if ci not in start_col_per_combo:
            continue
        present_y = [y for y in y_cols if any((x, y) in lookup for gx in x_groups.values() for x in gx)]
        sc = start_col_per_combo[ci]
        for jy, y in enumerate(present_y):
            rv = ""
            for gx in x_groups.values():
                for x in gx:
                    r = lookup.get((x, y))
                    if r is not None and r2_col in r.index:
                        v = r.get(r2_col, np.nan)
                        if not pd.isna(v):
                            rv = f"{v:.3f}"; break
                if rv != "": break
            cell = ws.cell(row_idx, sc + jy, rv)
            cell.alignment = C; cell.font = BOT_FONT
            cell.fill = BOT_FILL; cell.border = MED
    row_idx += 1

    max_col = max(
        start_col_per_combo[ci] + sum(1 for y in y_cols if any((x, y) in lk for gx in x_groups.values() for x in gx)) - 1
        for ci, (lk, _, _) in enumerate(all_lookups) if ci in start_col_per_combo
    ) if start_col_per_combo else 2
    c = ws.cell(row_idx, 1, NOTE_TEXT)
    c.font = Font(size=8, italic=True, color="595959"); c.alignment = L
    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=max_col)

    ws.column_dimensions[get_column_letter(1)].width = 20
    for col in range(2, max_col + 2):
        ws.column_dimensions[get_column_letter(col)].width = 10


def write_mod_sheet(ws, all_lookups, y_cols, x_groups,
                    coef_x, tstat_x, pval_x,
                    coef_mod, pval_mod,
                    coef_int, tstat_int, pval_int,
                    n_ctrl_col, r2_col, mod_label, sheet_title):
    ws.title = sheet_title[:31]

    start_col_per_combo, _ = _write_double_header(ws, y_cols, all_lookups)

    row_idx = 3
    for grp_lbl, x_list in x_groups.items():
        active = []
        for x in x_list:
            for lookup, ci, cl in all_lookups:
                if any((x, y) in lookup for y in y_cols):
                    active.append(x); break
        if not active:
            continue

        c = ws.cell(row_idx, 1, f"-- {grp_lbl.upper()} --")
        c.font = XGRP_FONT; c.fill = XGRP_FILL; c.alignment = L
        row_idx += 1

        for x in active:
            x_lbl = X_LABEL_MAP.get(x, x)

            ws.cell(row_idx,     1, x_lbl).alignment = L
            ws.cell(row_idx,     1).font = Font(size=9)
            ws.cell(row_idx + 1, 1, "")
            ws.cell(row_idx + 2, 1, f"  {mod_label}").alignment = L
            ws.cell(row_idx + 2, 1).font = Font(size=9, italic=True)
            ws.cell(row_idx + 3, 1, f"  X\u00d7{mod_label}").alignment = L
            ws.cell(row_idx + 3, 1).font = Font(size=9, bold=True, color="C00000")
            ws.cell(row_idx + 4, 1, "")

            for ci, (lookup, combo_idx, _) in enumerate(all_lookups):
                if ci not in start_col_per_combo:
                    continue
                present_y = [y for y in y_cols if any((x2, y) in lookup for gx in x_groups.values() for x2 in gx)]
                sc = start_col_per_combo[ci]
                tstat_fill = TSTAT_FILLS[combo_idx % 4]

                for jy, y in enumerate(present_y):
                    r = lookup.get((x, y))
                    ci_col = sc + jy

                    def cv(rr, col): return rr.get(col, np.nan) if rr is not None else np.nan

                    cell = ws.cell(row_idx, ci_col,
                                   fmt_coef(cv(r, coef_x), cv(r, pval_x)))
                    cell.alignment = C; cell.font = Font(size=9)
                    cell = ws.cell(row_idx + 1, ci_col,
                                   fmt_tstat(cv(r, tstat_x)))
                    cell.alignment = C; cell.font = Font(size=8, color="595959")
                    cell.fill = tstat_fill
                    cell = ws.cell(row_idx + 2, ci_col,
                                   fmt_coef(cv(r, coef_mod), cv(r, pval_mod)))
                    cell.alignment = C; cell.font = Font(size=9, italic=True)
                    cell = ws.cell(row_idx + 3, ci_col,
                                   fmt_coef(cv(r, coef_int), cv(r, pval_int)))
                    cell.alignment = C; cell.font = Font(size=9, bold=True, color="C00000")
                    cell = ws.cell(row_idx + 4, ci_col,
                                   fmt_tstat(cv(r, tstat_int)))
                    cell.alignment = C; cell.font = Font(size=8, color="C00000")
                    cell.fill = INT_FILL

            row_idx += 5

    # N
    c = ws.cell(row_idx, 1, "N")
    c.font = BOT_FONT; c.fill = BOT_FILL; c.alignment = L; c.border = THIN
    for ci, (lookup, combo_idx, _) in enumerate(all_lookups):
        if ci not in start_col_per_combo:
            continue
        present_y = [y for y in y_cols if any((x, y) in lookup for gx in x_groups.values() for x in gx)]
        sc = start_col_per_combo[ci]
        for jy, y in enumerate(present_y):
            n_val = ""
            for gx in x_groups.values():
                for x in gx:
                    r = lookup.get((x, y))
                    if r is not None:
                        nv = r.get(n_ctrl_col, np.nan)
                        if not pd.isna(nv):
                            n_val = int(nv); break
                if n_val != "": break
            cell = ws.cell(row_idx, sc + jy, n_val)
            cell.alignment = C; cell.font = BOT_FONT
            cell.fill = BOT_FILL; cell.border = THIN
    row_idx += 1

    # R²
    c = ws.cell(row_idx, 1, "R\u00b2(ctrl)")
    c.font = BOT_FONT; c.fill = BOT_FILL; c.alignment = L; c.border = MED
    for ci, (lookup, combo_idx, _) in enumerate(all_lookups):
        if ci not in start_col_per_combo:
            continue
        present_y = [y for y in y_cols if any((x, y) in lookup for gx in x_groups.values() for x in gx)]
        sc = start_col_per_combo[ci]
        for jy, y in enumerate(present_y):
            rv = ""
            for gx in x_groups.values():
                for x in gx:
                    r = lookup.get((x, y))
                    if r is not None and r2_col in r.index:
                        v = r.get(r2_col, np.nan)
                        if not pd.isna(v):
                            rv = f"{v:.3f}"; break
                if rv != "": break
            cell = ws.cell(row_idx, sc + jy, rv)
            cell.alignment = C; cell.font = BOT_FONT
            cell.fill = BOT_FILL; cell.border = MED
    row_idx += 1

    max_col = max(
        start_col_per_combo[ci] + sum(1 for y in y_cols if any((x, y) in lk for gx in x_groups.values() for x in gx)) - 1
        for ci, (lk, _, _) in enumerate(all_lookups) if ci in start_col_per_combo
    ) if start_col_per_combo else 2
    c = ws.cell(row_idx, 1, NOTE_TEXT)
    c.font = Font(size=8, italic=True, color="595959"); c.alignment = L
    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=max_col)

    ws.column_dimensions[get_column_letter(1)].width = 22
    for col in range(2, max_col + 2):
        ws.column_dimensions[get_column_letter(col)].width = 10


def get_mod_cols(df):
    suffix = None
    for c in df.columns:
        if c.startswith("coef_interact_") and not c.endswith("_ctrl"):
            suffix = c.replace("coef_interact_", ""); break
    if suffix is None:
        suffix = "mkt"
    coef_x   = "coef_ctrl";  tstat_x = "tstat_ctrl";  pval_x = "pvalue_ctrl"
    coef_mod = f"coef_{suffix}_ctrl" if f"coef_{suffix}_ctrl" in df.columns else "coef_mkt_ctrl"
    pval_mod = f"pval_{suffix}_ctrl" if f"pval_{suffix}_ctrl" in df.columns else "pval_mkt_ctrl"
    coef_int  = f"coef_interact_{suffix}_ctrl"  if f"coef_interact_{suffix}_ctrl"  in df.columns else "coef_interact_ctrl"
    tstat_int = f"tstat_interact_{suffix}_ctrl" if f"tstat_interact_{suffix}_ctrl" in df.columns else "tstat_interact_ctrl"
    pval_int  = f"pvalue_interact_{suffix}_ctrl" if f"pvalue_interact_{suffix}_ctrl" in df.columns else "pvalue_interact_ctrl"
    return coef_x, tstat_x, pval_x, coef_mod, pval_mod, coef_int, tstat_int, pval_int


# ─── Base filename fragment shared by all every-panel CSVs ───────────────────
_E = "every_after_end_rc1_rc2_ic_pc_yfe_ife_w99"
_S = "every_after_start_rc1_rc2_ic_pc_yfe_ife_w99"

# ─── Extended Y label map for top-rival suffixes ──────────────────────────────
for _k, _v in list(Y_LABEL_MAP.items()):
    for _top in ("_top1", "_top3", "_top5", "_top10"):
        Y_LABEL_MAP[_k + _top] = _v + f"\n({_top[1:]})"

# ─── Extended Y orders (original + top1 + top3 + top5 + top10) ───────────────
Y_ORDER_END_EXT   = (Y_ORDER_END
                     + [y + "_top1"  for y in Y_ORDER_END]
                     + [y + "_top3"  for y in Y_ORDER_END]
                     + [y + "_top5"  for y in Y_ORDER_END]
                     + [y + "_top10" for y in Y_ORDER_END])
Y_ORDER_START_EXT = (Y_ORDER_START
                     + [y + "_top1"  for y in Y_ORDER_START]
                     + [y + "_top3"  for y in Y_ORDER_START]
                     + [y + "_top5"  for y in Y_ORDER_START]
                     + [y + "_top10" for y in Y_ORDER_START])

# ─── Load data ────────────────────────────────────────────────────────────────
print("Loading data...")
df_end   = pd.read_csv(REG_DIR / f"reg_bivariate_grouped_{_E}.csv")
df_start = pd.read_csv(REG_DIR / f"reg_bivariate_grouped_{_S}.csv")
df_end_top1   = pd.read_csv(REG_DIR / f"reg_bivariate_grouped_{_E}_top1.csv")
df_end_top3   = pd.read_csv(REG_DIR / f"reg_bivariate_grouped_{_E}_top3.csv")
df_end_top5   = pd.read_csv(REG_DIR / f"reg_bivariate_grouped_{_E}_top5.csv")
df_start_top1 = pd.read_csv(REG_DIR / f"reg_bivariate_grouped_{_S}_top1.csv")
df_start_top3 = pd.read_csv(REG_DIR / f"reg_bivariate_grouped_{_S}_top3.csv")
df_start_top5  = pd.read_csv(REG_DIR / f"reg_bivariate_grouped_{_S}_top5.csv")
df_end_top10   = pd.read_csv(REG_DIR / f"reg_bivariate_grouped_{_E}_top10.csv")
df_start_top10 = pd.read_csv(REG_DIR / f"reg_bivariate_grouped_{_S}_top10.csv")
df_end_mkt    = pd.read_csv(REG_DIR / f"reg_bivariate_grouped_every_after_end_rc1_rc2_ic_pc_yfe_ife_mkt_w99.csv")
df_start_mkt  = pd.read_csv(REG_DIR / f"reg_bivariate_grouped_every_after_start_rc1_rc2_ic_pc_yfe_ife_mkt_w99.csv")
df_end_comp   = pd.read_csv(REG_DIR / f"reg_bivariate_grouped_{_E}_comp_verbal.csv")
df_start_comp = pd.read_csv(REG_DIR / f"reg_bivariate_grouped_{_S}_comp_verbal.csv")
df_end_pros   = pd.read_csv(REG_DIR / f"reg_bivariate_grouped_{_E}_pros_verbal.csv")
df_start_pros = pd.read_csv(REG_DIR / f"reg_bivariate_grouped_{_S}_pros_verbal.csv")
df_end_acq    = pd.read_csv(REG_DIR / f"reg_bivariate_grouped_{_E}_a_comp_qa.csv")
df_end_qcq    = pd.read_csv(REG_DIR / f"reg_bivariate_grouped_{_E}_q_comp_qa.csv")
df_end_apq    = pd.read_csv(REG_DIR / f"reg_bivariate_grouped_{_E}_a_pros_qa.csv")
df_end_qpq    = pd.read_csv(REG_DIR / f"reg_bivariate_grouped_{_E}_q_pros_qa.csv")
df_start_acq  = pd.read_csv(REG_DIR / f"reg_bivariate_grouped_{_S}_a_comp_qa.csv")
df_start_qcq  = pd.read_csv(REG_DIR / f"reg_bivariate_grouped_{_S}_q_comp_qa.csv")
df_start_apq  = pd.read_csv(REG_DIR / f"reg_bivariate_grouped_{_S}_a_pros_qa.csv")
df_start_qpq  = pd.read_csv(REG_DIR / f"reg_bivariate_grouped_{_S}_q_pros_qa.csv")
print("Loaded.")

# ─── Build workbook ───────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
wb.remove(wb.active)

def build_all_lookups(df, extra_dfs=None):
    """extra_dfs: list of (dataframe, y_suffix) pairs merged into each combo lookup."""
    result = []
    for ci, (sess, grp, label) in enumerate(ALL_COMBOS):
        lk = _build_lookup(df, sess, grp)
        if extra_dfs:
            for extra_df, suffix in extra_dfs:
                lk.update(_build_lookup_suffixed(extra_df, sess, grp, suffix))
        result.append((lk, ci, label))
    return result

MAIN_PANELS = [
    ("A1_End_CAR",   df_end,   Y_ORDER_END_EXT,
     [(df_end_top1, "_top1"), (df_end_top3, "_top3"), (df_end_top5, "_top5"), (df_end_top10, "_top10")],
     "coef_ctrl", "tstat_ctrl", "pvalue_ctrl"),
    ("A2_Start_CAR", df_start, Y_ORDER_START_EXT,
     [(df_start_top1, "_top1"), (df_start_top3, "_top3"), (df_start_top5, "_top5"), (df_start_top10, "_top10")],
     "coef_ctrl", "tstat_ctrl", "pvalue_ctrl"),
]

for sheet_name, df_, y_order, extra_dfs, cc, tc, pc in MAIN_PANELS:
    ws = wb.create_sheet(sheet_name)
    all_lookups = build_all_lookups(df_, extra_dfs)
    write_sheet(ws, all_lookups, y_order, X_GROUPS,
                coef_col=cc, tstat_col=tc, pval_col=pc,
                n_ctrl_col="n_obs_ctrl", r2_col="r2_ctrl",
                sheet_title=sheet_name)
    print(f"  Written: {sheet_name}")

MOD_PANELS = [
    ("B1_End_Mkt",     df_end_mkt,    Y_ORDER_END,   "Mkt"),
    ("B2_Str_Mkt",     df_start_mkt,  Y_ORDER_START, "Mkt"),
    ("C1_End_CompV",   df_end_comp,   Y_ORDER_END,   "CompVerb"),
    ("C2_Str_CompV",   df_start_comp, Y_ORDER_START, "CompVerb"),
    ("D1_End_ProsV",   df_end_pros,   Y_ORDER_END,   "ProsVerb"),
    ("D2_Str_ProsV",   df_start_pros, Y_ORDER_START, "ProsVerb"),
    ("E1_End_ACompQA", df_end_acq,    Y_ORDER_END,   "A_CompQA"),
    ("E2_End_QCompQA", df_end_qcq,    Y_ORDER_END,   "Q_CompQA"),
    ("E3_End_AProsQA", df_end_apq,    Y_ORDER_END,   "A_ProsQA"),
    ("E4_End_QProsQA", df_end_qpq,    Y_ORDER_END,   "Q_ProsQA"),
    ("F1_Str_ACompQA", df_start_acq,  Y_ORDER_START, "A_CompQA"),
    ("F2_Str_QCompQA", df_start_qcq,  Y_ORDER_START, "Q_CompQA"),
    ("F3_Str_AProsQA", df_start_apq,  Y_ORDER_START, "A_ProsQA"),
    ("F4_Str_QProsQA", df_start_qpq,  Y_ORDER_START, "Q_ProsQA"),
]

for sheet_name, df_, y_order, mod_label in MOD_PANELS:
    ws = wb.create_sheet(sheet_name)
    all_lookups = build_all_lookups(df_)
    cx, tx, px, cm, pm, ci, ti, pi = get_mod_cols(df_)
    write_mod_sheet(ws, all_lookups, y_order, X_GROUPS,
                    coef_x=cx, tstat_x=tx, pval_x=px,
                    coef_mod=cm, pval_mod=pm,
                    coef_int=ci, tstat_int=ti, pval_int=pi,
                    n_ctrl_col="n_obs_ctrl", r2_col="r2_ctrl",
                    mod_label=mod_label, sheet_title=sheet_name)
    print(f"  Written: {sheet_name}")

wb.save(OUT_FILE)
import os
print(f"\nSaved: {OUT_FILE}  ({os.path.getsize(OUT_FILE)//1024} KB)")
print(f"Total sheets: {len(wb.worksheets)}")
